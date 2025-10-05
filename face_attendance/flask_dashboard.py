# flask_dashboard_enhanced.py
# Enhanced admin dashboard with analytics, charts, exports, Slack, and optional camera stream

import os
import io
import time
import threading
import datetime as dt
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

from flask import (
    Flask, request, redirect, url_for, render_template, send_file,
    jsonify, flash, Response
)
from werkzeug.utils import secure_filename

from sqlalchemy import func
from sqlalchemy.exc import OperationalError

from db import SessionLocal, User, Attendance, init_db

# ---------------- Config ----------------
APP_TITLE = os.getenv("APP_TITLE", "Attendance Admin")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")
UPLOAD_DIR = os.path.abspath(os.getenv("UPLOAD_DIR", os.path.join(BASE_DIR, "uploads")))
SEED_DIR = os.path.abspath(os.getenv("SEED_DIR", os.path.join(BASE_DIR, "seed_images")))
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(SEED_DIR, exist_ok=True)

# Slack (optional)
SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL")
SLACK_CHANNEL = os.getenv("SLACK_CHANNEL")

# Camera (optional)
ENABLE_CAMERA_STREAM = os.getenv("ENABLE_CAMERA_STREAM", "1") == "1"
CAMERA_INDEX = int(os.getenv("CAMERA_INDEX", "0"))

# Gates & late cutoffs (edit freely)
GATES: Dict[int, Tuple[str, str, str]] = {
    1: ("Morning Start", "08:00", "09:30"),
    2: ("Lunch Out", "12:30", "13:30"),
    3: ("Lunch In", "13:00", "14:30"),
    4: ("Evening End", "17:00", "19:30"),
}
LATE_AT: Dict[int, str] = {1: "08:00", 2: "12:35", 3: "13:05", 4: "17:05"}

# Upload safety
ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}

app = Flask(__name__, template_folder="template", static_folder="static")

app.secret_key = os.getenv("FLASK_SECRET", "dev-secret")  # replace in prod
app.jinja_env.auto_reload = True

# ---------------- Utils -----------------
def today_str() -> str:
    return dt.date.today().isoformat()

def parse_hhmm(s: str) -> dt.time:
    h, m = map(int, s.split(":")[:2])
    return dt.time(hour=h, minute=m)

def is_late(gate_id: int, tstr: Optional[str]) -> bool:
    if not tstr:
        return False
    late_cut = parse_hhmm(LATE_AT.get(gate_id, "23:59"))
    t = dt.datetime.strptime(tstr, "%H:%M:%S").time() if ":" in tstr else parse_hhmm(tstr)
    return t > late_cut

def quarter_for(d: dt.date) -> str:
    q = (d.month - 1) // 3 + 1
    return f"{d.year}{q}"

def week_bounds(any_day: dt.date) -> Tuple[dt.date, dt.date]:
    start = any_day - dt.timedelta(days=any_day.weekday())
    end = start + dt.timedelta(days=6)
    return start, end

def month_bounds(any_day: dt.date) -> Tuple[dt.date, dt.date]:
    start = any_day.replace(day=1)
    next_month = start.replace(day=28) + dt.timedelta(days=4)
    end = next_month - dt.timedelta(days=next_month.day)
    return start, end

def slack_post(text: str):
    if not SLACK_WEBHOOK_URL:
        return
    try:
        import requests
        payload = {"text": text}
        if SLACK_CHANNEL:
            payload["channel"] = SLACK_CHANNEL
        requests.post(SLACK_WEBHOOK_URL, json=payload, timeout=5)
    except Exception:
        pass

# ------------- Metrics ------------------
@dataclass
class DashboardMetrics:
    total_users: int
    present_today: int
    absent_today: int
    late_today: int
    on_time_today: int
    weekly_present: List[int]
    weekly_absent: List[int]
    weekly_late: List[int]
    weekly_labels: List[str]
    department_stats: Dict[str, Dict[str, int]]
    late_employees: List[Dict]
    absent_employees: List[str]

def compute_dashboard_metrics(sess) -> DashboardMetrics:
    total_users = sess.query(User).count()
    today = today_str()

    # Today's stats
    today_records = sess.query(Attendance).filter(Attendance.date == today).all()
    present_today = len([r for r in today_records if r.checkin1_time])
    absent_today = max(0, total_users - present_today)
    late_today = len([r for r in today_records if r.checkin1_time and is_late(1, r.checkin1_time)])
    on_time_today = max(0, present_today - late_today)

    # Weekly (last 7 days)
    start_week = dt.date.today() - dt.timedelta(days=6)
    weekly_present, weekly_absent, weekly_late, weekly_labels = [], [], [], []
    for i in range(7):
        day = start_week + dt.timedelta(days=i)
        s = day.isoformat()
        recs = sess.query(Attendance).filter(Attendance.date == s).all()
        present = len([r for r in recs if r.checkin1_time])
        late = len([r for r in recs if r.checkin1_time and is_late(1, r.checkin1_time)])
        weekly_present.append(present)
        weekly_absent.append(max(0, total_users - present))
        weekly_late.append(late)
        weekly_labels.append(day.strftime("%a %m/%d"))

    # Late employees (today)
    late_employees = []
    for r in today_records:
        if r.checkin1_time and is_late(1, r.checkin1_time):
            user = sess.get(User, r.user_id)
            late_employees.append({"name": user.name, "time": r.checkin1_time, "user_id": user.user_id})

    # Absent employees (today)
    present_ids = {r.user_id for r in today_records if r.checkin1_time}
    absent_employees = [u.name for u in sess.query(User).all() if u.user_id not in present_ids]

    department_stats = {"All": {"present": present_today, "absent": absent_today, "late": late_today}}

    return DashboardMetrics(
        total_users, present_today, absent_today, late_today, on_time_today,
        weekly_present, weekly_absent, weekly_late, weekly_labels,
        department_stats, late_employees, absent_employees
    )

# ------------- Pages / APIs ------------
@app.get("/", endpoint="home")
def home():
    sess = SessionLocal()
    try:
        metrics = compute_dashboard_metrics(sess)
    finally:
        sess.close()
    return render_template("dashboard_enhanced.html",
                           title=APP_TITLE, today=dt.date.today(), metrics=metrics, gates=GATES)

@app.get("/users", endpoint="users_page")
def users_page():
    sess = SessionLocal()
    try:
        users = sess.query(User).order_by(User.user_id).all()
    finally:
        sess.close()
    return render_template("users.html", title=APP_TITLE, users=users)

@app.post("/users/add")
def users_add():
    name = (request.form.get("name") or "").strip()
    if not name:
        flash("Name is required", "error")
        return redirect(url_for("users_page"))

    file = request.files.get("photo")
    if not (file and file.filename):
        flash("Photo is required to enroll", "error")
        return redirect(url_for("users_page"))

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXT:
        flash("Unsupported image format", "error")
        return redirect(url_for("users_page"))

    safe_name = secure_filename(name) or "user"
    filepath = os.path.join(SEED_DIR, f"{safe_name}{ext}")
    file.save(filepath)

    # Enroll with DeepFace (no 'model' kwarg—works across versions)
    try:
        import numpy as np
        from PIL import Image, ImageOps
        import cv2
        from deepface import DeepFace
        import json

        def read_bgr(pth: str):
            im = Image.open(pth)
            im = ImageOps.exif_transpose(im)
            rgb = im.convert("RGB")
            arr = np.array(rgb)
            return cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)

        model_name = os.getenv("MODEL_NAME", "Facenet")  # fast & compatible
        reps = DeepFace.represent(
            img_path=read_bgr(filepath),
            model_name=model_name,
            detector_backend=os.getenv("DETECTOR", "opencv"),
            enforce_detection=True,
            align=True,
        )
        if not reps:
            flash("Could not detect a face clearly in that photo.", "error")
            return redirect(url_for("users_page"))

        emb = reps[0]["embedding"]
        sess = SessionLocal()
        try:
            u = sess.query(User).filter_by(name=name).one_or_none()
            if u:
                u.face_embedding = json.dumps(emb)
            else:
                u = User(name=name, face_embedding=json.dumps(emb))
                sess.add(u)
            sess.commit()
        finally:
            sess.close()

        flash(f"Enrolled {name}", "success")
    except Exception as e:
        flash(f"Enroll failed: {e}", "error")

    return redirect(url_for("users_page"))

@app.post("/users/delete/<int:user_id>")
def users_delete(user_id: int):
    sess = SessionLocal()
    try:
        u = sess.get(User, user_id)
        if not u:
            flash("User not found", "error")
        else:
            sess.delete(u)
            sess.commit()
            flash("User deleted", "success")
    finally:
        sess.close()
    return redirect(url_for("users_page"))

@app.get("/attendance", endpoint="attendance_page")
def attendance_page():
    d = request.args.get("date") or today_str()
    sess = SessionLocal()
    try:
        rows = (
            sess.query(Attendance, User)
            .join(User, Attendance.user_id == User.user_id)
            .filter(Attendance.date == d)
            .order_by(User.name)
            .all()
        )

        def badge(val: Optional[str], gate_id: int) -> Tuple[str, str]:
            if not val:
                return ("-", "muted")
            return (val, "late" if is_late(gate_id, val) else "ok")

        data = [{
            "user_id": user.user_id,
            "name": user.name,
            "g1": badge(att.checkin1_time, 1),
            "g2": badge(att.checkin2_time, 2),
            "g3": badge(att.checkin3_time, 3),
            "g4": badge(att.checkin4_time, 4),
            "quarter": att.quarter_id,
            "date": att.date,
        } for att, user in rows]
    finally:
        sess.close()

    return render_template("attendance.html", title=APP_TITLE, date=d, rows=data, gates=GATES)

# ---------- Analytics ----------
@app.get("/analytics", endpoint="analytics_page")
def analytics_page():
    period = request.args.get("period", "week")  # week|month
    if period == "month":
        start, end = month_bounds(dt.date.today())
    else:
        start, end = week_bounds(dt.date.today())

    sess = SessionLocal()
    try:
        records = (
            sess.query(Attendance)
            .filter(Attendance.date >= start.isoformat(), Attendance.date <= end.isoformat())
            .all()
        )
        total_users = sess.query(User).count()

        daily_stats = {}
        current = start
        while current <= end:
            s = current.isoformat()
            day_recs = [r for r in records if r.date == s]
            present = len([r for r in day_recs if r.checkin1_time])
            late = len([r for r in day_recs if r.checkin1_time and is_late(1, r.checkin1_time)])
            daily_stats[s] = {
                "present": present,
                "absent": max(0, total_users - present),
                "late": late,
                "on_time": max(0, present - late),
            }
            current += dt.timedelta(days=1)
    finally:
        sess.close()

    return render_template("analytics.html", title=APP_TITLE, period=period,
                           start_date=start, end_date=end, daily_stats=daily_stats, total_users=total_users)

@app.get("/api/analytics/weekly-absent")
def api_weekly_absent():
    sess = SessionLocal()
    try:
        start_week = dt.date.today() - dt.timedelta(days=6)
        total_users = sess.query(User).count()
        data = []
        for i in range(7):
            day = start_week + dt.timedelta(days=i)
            s = day.isoformat()
            present = sess.query(Attendance).filter(
                Attendance.date == s, Attendance.checkin1_time.isnot(None)
            ).count()
            data.append({
                "date": day.strftime("%Y-%m-%d"),
                "label": day.strftime("%a %m/%d"),
                "absent": max(0, total_users - present),
                "present": present
            })
    finally:
        sess.close()
    return jsonify(data)

@app.get("/api/analytics/monthly-trends")
def api_monthly_trends():
    sess = SessionLocal()
    try:
        start, end = month_bounds(dt.date.today())
        total_users = sess.query(User).count()
        data = []
        current = start
        while current <= end:
            s = current.isoformat()
            recs = sess.query(Attendance).filter(Attendance.date == s).all()
            present = len([r for r in recs if r.checkin1_time])
            late = len([r for r in recs if r.checkin1_time and is_late(1, r.checkin1_time)])
            data.append({
                "date": s,
                "present": present,
                "absent": max(0, total_users - present),
                "late": late,
                "on_time": max(0, present - late)
            })
            current += dt.timedelta(days=1)
    finally:
        sess.close()
    return jsonify(data)

# ---------- Exports ----------
@app.get("/export/weekly-excel")
def export_weekly_excel():
    try:
        base = dt.datetime.strptime(request.args.get("week_of", today_str()), "%Y-%m-%d").date()
    except Exception:
        base = dt.date.today()
    start, end = week_bounds(base)

    sess = SessionLocal()
    try:
        all_users = sess.query(User).order_by(User.name).all()
        recs = (
            sess.query(Attendance)
            .filter(Attendance.date >= start.isoformat(), Attendance.date <= end.isoformat())
            .all()
        )
        att_map = {(r.user_id, r.date): r for r in recs}
    finally:
        sess.close()

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Attendance"

    headers = ["Employee"] + [(start + dt.timedelta(days=i)).strftime("%a %m/%d") for i in range(7)]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    absent_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    late_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    present_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

    for user in all_users:
        row_vals = [user.name]
        for i in range(7):
            day = start + dt.timedelta(days=i)
            s = day.isoformat()
            att = att_map.get((user.user_id, s))
            if att and att.checkin1_time:
                status = "LATE" if is_late(1, att.checkin1_time) else "Present"
                row_vals.append(f"{att.checkin1_time} ({status})")
            else:
                row_vals.append("ABSENT")
        ws.append(row_vals)
        r = ws.max_row
        for c, val in enumerate(row_vals[1:], start=2):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            s = str(val)
            if "ABSENT" in s:
                cell.fill = absent_fill
            elif "LATE" in s:
                cell.fill = late_fill
            elif "Present" in s:
                cell.fill = present_fill

    # autosize
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fn = f"attendance_week_{start.isoformat()}_{end.isoformat()}.xlsx"
    return send_file(bio, as_attachment=True, download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/export/absent-report")
def export_absent_report():
    period = request.args.get("period", "week")
    if period == "month":
        start, end = month_bounds(dt.date.today())
    else:
        start, end = week_bounds(dt.date.today())

    sess = SessionLocal()
    try:
        all_users = sess.query(User).order_by(User.name).all()
    finally:
        sess.close()

    import pandas as pd
    rows = []
    for u in all_users:
        absent_days, total_days = [], 0
        cur = start
        while cur <= end:
            if cur.weekday() < 5:
                total_days += 1
                s = cur.isoformat()
                sess = SessionLocal()
                try:
                    att = sess.query(Attendance).filter(
                        Attendance.user_id == u.user_id, Attendance.date == s
                    ).first()
                finally:
                    sess.close()
                if not att or not att.checkin1_time:
                    absent_days.append(s)
            cur += dt.timedelta(days=1)
        rate = f"{(len(absent_days)/total_days*100):.1f}%" if total_days else "0%"
        rows.append({
            "Employee": u.name,
            "Total Working Days": total_days,
            "Days Absent": len(absent_days),
            "Absence Rate": rate,
            "Absent Dates": ", ".join(absent_days) if absent_days else "None"
        })

    df = pd.DataFrame(rows)
    bio = io.BytesIO()
    df.to_excel(bio, index=False, engine="openpyxl")
    bio.seek(0)
    fn = f"absent_report_{start.isoformat()}_{end.isoformat()}.xlsx"
    return send_file(bio, as_attachment=True, download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------- Camera stream (optional) ----------
def gen_camera():
    import cv2
    cap = cv2.VideoCapture(CAMERA_INDEX)
    if not cap.isOpened():
        return
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    cap.set(cv2.CAP_PROP_FPS, 30)
    try:
        while True:
            ok, frame = cap.read()
            if not ok:
                break
            ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cv2.putText(frame, ts, (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,255), 2)
            ok, buf = cv2.imencode(".jpg", frame)
            if not ok:
                continue
            yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + bytearray(buf) + b"\r\n")
            time.sleep(0.03)
    finally:
        cap.release()

@app.get("/video_feed")
def video_feed():
    if not ENABLE_CAMERA_STREAM:
        return "Disabled", 404
    return Response(gen_camera(), mimetype="multipart/x-mixed-replace; boundary=frame")

# ---------- REST for engine to log ----------
@app.post("/api/attendance/log")
def api_attendance_log():
    data = request.get_json(force=True, silent=True) or {}
    user_id = data.get("user_id")
    gate_id = data.get("gate_id")
    t = data.get("time")
    d = data.get("date") or today_str()
    if not (user_id and gate_id and t):
        return jsonify(ok=False, error="user_id, gate_id, time required"), 400

    sess = SessionLocal()
    try:
        att = sess.query(Attendance).filter(Attendance.user_id == user_id, Attendance.date == d).one_or_none()
        if not att:
            att = Attendance(user_id=user_id, date=d, quarter_id=quarter_for(dt.datetime.strptime(d, "%Y-%m-%d").date()))
            sess.add(att)
            sess.flush()

        updated = False
        if gate_id == 1 and (not att.checkin1_time or att.checkin1_time != t):
            att.checkin1_time = t; updated = True
        elif gate_id == 2 and (not att.checkin2_time or att.checkin2_time != t):
            att.checkin2_time = t; updated = True
        elif gate_id == 3 and (not att.checkin3_time or att.checkin3_time != t):
            att.checkin3_time = t; updated = True
        elif gate_id == 4 and (not att.checkin4_time or att.checkin4_time != t):
            att.checkin4_time = t; updated = True

        sess.commit()

        if updated:
            user = sess.get(User, user_id)
            status = "LATE" if is_late(gate_id, t) else "On-time"
            slack_post(f":white_check_mark: {user.name} → Gate {gate_id} at {t} ({status})")
    finally:
        sess.close()

    return jsonify(ok=True, updated=updated)

# ---------- Background watcher ----------
class AttendanceWatcher(threading.Thread):
    def __init__(self, interval_sec: int = 10):
        super().__init__(daemon=True)
        self.interval = interval_sec
        self._stop = threading.Event()
        self._seen = {}

    def run(self):
        while not self._stop.is_set():
            try:
                sess = SessionLocal()
                rows = sess.query(Attendance).filter(Attendance.date == today_str()).all()
                for r in rows:
                    key = (r.user_id, r.date)
                    curr = (r.checkin1_time, r.checkin2_time, r.checkin3_time, r.checkin4_time)
                    prev = self._seen.get(key)
                    if prev and prev != curr:
                        for gate_id, idx in {1:0, 2:1, 3:2, 4:3}.items():
                            if (prev[idx] or "") != (curr[idx] or "") and curr[idx]:
                                user = sess.get(User, r.user_id)
                                status = "LATE" if is_late(gate_id, curr[idx]) else "On-time"
                                slack_post(f":alarm_clock: {user.name} marked Gate {gate_id} at {curr[idx]} — {status}")
                    self._seen[key] = curr
                sess.close()
            except Exception:
                pass
            self._stop.wait(self.interval)

    def stop(self):
        self._stop.set()

watcher: Optional[AttendanceWatcher] = None

def start_backgrounds():
    global watcher
    if watcher is None:
        watcher = AttendanceWatcher(interval_sec=10)
        watcher.start()

def stop_backgrounds():
    global watcher
    if watcher:
        watcher.stop()
        watcher = None

@app.context_processor
def inject_globals():
    return {"APP_TITLE": APP_TITLE, "now": dt.datetime.now, "gates": GATES, "url_for": url_for}

@app.get("/health")
def health():
    # Quick DB ping
    ok = True
    try:
        s = SessionLocal(); s.execute(func.now()); s.close()
    except Exception:
        ok = False
    return jsonify(ok=ok, time=dt.datetime.utcnow().isoformat())

if __name__ == "__main__":
    # If MySQL is down, let db.py fallback to SQLite if you added that logic
    try:
        init_db()
    except OperationalError as e:
        print("DB init failed:", e)
        raise
    start_backgrounds()
    try:
        app.run(debug=True, threaded=True)
    finally:
        stop_backgrounds()
